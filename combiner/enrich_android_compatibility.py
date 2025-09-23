import csv
import os
import re
import sys
import time
import hashlib
import argparse
from typing import Dict, List, Optional, Tuple

try:
    from bs4 import BeautifulSoup
    HAVE_BS4 = True
except Exception:
    HAVE_BS4 = False
    BeautifulSoup = None  # type: ignore

try:
    import requests
    from requests.adapters import HTTPAdapter
    from urllib3.util.retry import Retry
    HAVE_REQ = True
except Exception:
    HAVE_REQ = False

# Optional: headless rendering for dynamic Google Play content
try:
    from playwright.sync_api import sync_playwright  # type: ignore
    HAVE_PLAYWRIGHT = True
except Exception:
    HAVE_PLAYWRIGHT = False

try:
    import openpyxl
    from openpyxl.worksheet.table import Table, TableStyleInfo
    HAVE_OPENPYXL = True
except Exception:
    HAVE_OPENPYXL = False


PLAY_PREFIX = "https://play.google.com"

# Canonical tokens -> output column names in sheet
COMPAT_TOKENS: Dict[str, List[str]] = {
    "mobile": ["mobile", "phone"],
    "tablet": ["tablet"],
    "chromebook": ["chromebook"],
    "tv": ["tv", "android tv"],
    "watch": ["watch", "wear", "wear os", "wearos"],
    "car": ["car", "android auto", "auto", "automotive"],
}

DEFAULT_COL_MAP = {
    "mobile": "Mobile",
    "tablet": "Tablet",
    "chromebook": "Chromebook",
    "tv": "TV",
    "watch": "Watch",
    "car": "Car",
}


def ensure_deps():
    if not HAVE_BS4:
        sys.stderr.write("BeautifulSoup4 not installed. Please install beautifulsoup4.\n")
        sys.exit(2)
    if not HAVE_REQ:
        sys.stderr.write("requests not installed. Please install requests.\n")
        sys.exit(2)


def make_session() -> 'requests.Session':
    s = requests.Session()
    retries = Retry(
        total=3,
        backoff_factor=1.0,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET"],
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retries)
    s.mount("http://", adapter)
    s.mount("https://", adapter)
    s.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/125.0.0.0 Safari/537.36"
        ),
        "Accept-Language": "en-US,en;q=0.9",
    })
    return s


def sha1(s: str) -> str:
    return hashlib.sha1(s.encode("utf-8")).hexdigest()


def load_page(url: str, session: 'requests.Session', cache_dir: str, delay: float = 0.6) -> Optional[str]:
    os.makedirs(cache_dir, exist_ok=True)
    key = sha1(url)
    p = os.path.join(cache_dir, f"{key}.html")
    if os.path.exists(p):
        try:
            with open(p, "r", encoding="utf-8", errors="ignore") as f:
                return f.read()
        except Exception:
            pass
    try:
        resp = session.get(url, timeout=20)
        if resp.status_code == 200 and resp.text:
            html = resp.text
            try:
                with open(p, "w", encoding="utf-8") as f:
                    f.write(html)
            except Exception:
                pass
            time.sleep(delay)
            return html
        else:
            if resp.status_code == 429:
                time.sleep(5)
            return None
    except Exception:
        return None


def load_page_rendered(url: str, cache_dir: str, delay: float = 0.6, timeout_ms: int = 8000) -> Optional[str]:
    """Render the page with a headless browser to capture dynamically populated
    'Ratings and reviews' content and device chips. Requires Playwright.
    Returns extracted section text or None on failure.
    """
    if not HAVE_PLAYWRIGHT:
        return None
    os.makedirs(cache_dir, exist_ok=True)
    key = sha1(url + "|render|ratings")
    p = os.path.join(cache_dir, f"{key}.txt")
    if os.path.exists(p):
        try:
            with open(p, "r", encoding="utf-8", errors="ignore") as f:
                return f.read()
        except Exception:
            pass
    try:
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True)
            ua = (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0.0.0 Safari/537.36"
            )
            context = browser.new_context(
                user_agent=ua,
                locale="en-US",
                extra_http_headers={"Accept-Language": "en-US,en;q=0.9"},
            )
            page = context.new_page()
            # Enforce en-US for more predictable strings
            target_url = url
            if "hl=" not in target_url:
                sep = "&" if "?" in target_url else "?"
                target_url = f"{target_url}{sep}hl=en&gl=us"
            page.goto(target_url, wait_until="domcontentloaded", timeout=timeout_ms)
            # The reviews section loads lazily; scroll to trigger data fetching
            for _ in range(5):
                page.mouse.wheel(0, 1200)
                page.wait_for_timeout(250)
            try:
                # Attempt to locate the Ratings and reviews container by aria-label
                container = page.locator('[aria-label*="Ratings and reviews" i]')
                if not container or container.count() == 0:
                    # Try heading then sibling container patterns
                    heading = page.get_by_role("heading", name=re.compile("ratings.*reviews", re.I))
                    if heading and heading.count() > 0:
                        heading.first.scroll_into_view_if_needed()
                        page.wait_for_timeout(500)
                        # Expand context around the heading
                        container = heading.first.locator("xpath=following-sibling::*[1]")
                # If still empty, try waiting briefly for network/render
                if (not container or container.count() == 0):
                    try:
                        page.wait_for_selector('[aria-label*="Ratings and reviews" i]', timeout=2000)
                        container = page.locator('[aria-label*="Ratings and reviews" i]')
                    except Exception:
                        pass
                if container and container.count() > 0:
                    container.first.scroll_into_view_if_needed()
                    page.wait_for_timeout(700)
            except Exception:
                pass

            # Collect visible text within the container plus device filter chips nearby
            text_parts: List[str] = []
            try:
                if container and container.count() > 0:
                    text_parts.append(container.first.inner_text(timeout=1000))
            except Exception:
                pass
            try:
                # Chips often appear as filter buttons under Ratings and reviews
                name_rx = re.compile(r"phone|tablet|chromebook|android\s*tv|tv|watch|wear(?:\s*os)?|car|android\s*auto|automotive|all devices", re.I)
                chips = page.get_by_role("button", name=name_rx)
                # Wait briefly for chips if not present initially
                if chips.count() == 0:
                    try:
                        page.wait_for_selector('button', timeout=1000)
                        chips = page.get_by_role("button", name=name_rx)
                    except Exception:
                        pass
                chip_texts = []
                n = min(max(chips.count(), 0), 30)
                for i in range(n):
                    try:
                        t = chips.nth(i).inner_text(timeout=500)
                        if t and name_rx.search(t):
                            chip_texts.append(t)
                    except Exception:
                        continue
                if chip_texts:
                    text_parts.append(" | ".join(sorted(set(chip_texts), key=lambda x: x.lower())))
            except Exception:
                pass

            # Fallback: whole page text is too heavy; avoid
            out_text = "\n".join([t for t in text_parts if t])
            try:
                with open(p, "w", encoding="utf-8") as f:
                    f.write(out_text)
            except Exception:
                pass
            page.close(); context.close(); browser.close()
            time.sleep(delay)
            return out_text
    except Exception:
        return None


class _PersistentRenderer:
    """Reuse a single Playwright browser/page across many URLs for speed."""
    def __init__(self, cache_dir: str, delay: float = 0.6, timeout_ms: int = 8000):
        self.cache_dir = cache_dir
        self.delay = delay
        self.timeout_ms = timeout_ms
        self._pw = None
        self._browser = None
        self._context = None
        self._page = None

    def start(self):
        if not HAVE_PLAYWRIGHT:
            return
        os.makedirs(self.cache_dir, exist_ok=True)
        from playwright.sync_api import sync_playwright  # type: ignore
        self._pw = sync_playwright().start()
        self._browser = self._pw.chromium.launch(headless=True)
        ua = (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/125.0.0.0 Safari/537.36"
        )
        self._context = self._browser.new_context(
            user_agent=ua,
            locale="en-US",
            extra_http_headers={"Accept-Language": "en-US,en;q=0.9"},
        )
        self._page = self._context.new_page()

    def stop(self):
        try:
            if self._page:
                self._page.close()
        except Exception:
            pass
        try:
            if self._context:
                self._context.close()
        except Exception:
            pass
        try:
            if self._browser:
                self._browser.close()
        except Exception:
            pass
        try:
            if self._pw:
                self._pw.stop()
        except Exception:
            pass
        self._pw = self._browser = self._context = self._page = None

    def fetch(self, url: str) -> Optional[str]:
        if not HAVE_PLAYWRIGHT or not self._page:
            return None
        # Cache key compatible with load_page_rendered but distinct enough
        key = sha1(url + "|render|ratings")
        p = os.path.join(self.cache_dir, f"{key}.txt")
        if os.path.exists(p):
            try:
                with open(p, "r", encoding="utf-8", errors="ignore") as f:
                    return f.read()
            except Exception:
                pass
        try:
            target_url = url
            if "hl=" not in target_url:
                sep = "&" if "?" in target_url else "?"
                target_url = f"{target_url}{sep}hl=en&gl=us"
            self._page.goto(target_url, wait_until="domcontentloaded", timeout=self.timeout_ms)
            for _ in range(5):
                self._page.mouse.wheel(0, 1200)
                self._page.wait_for_timeout(250)
            text_parts: List[str] = []
            container = None
            try:
                container = self._page.locator('[aria-label*="Ratings and reviews" i]')
                if not container or container.count() == 0:
                    heading = self._page.get_by_role("heading", name=re.compile("ratings.*reviews", re.I))
                    if heading and heading.count() > 0:
                        heading.first.scroll_into_view_if_needed()
                        self._page.wait_for_timeout(500)
                        container = heading.first.locator("xpath=following-sibling::*[1]")
                if (not container or container.count() == 0):
                    try:
                        self._page.wait_for_selector('[aria-label*="Ratings and reviews" i]', timeout=2000)
                        container = self._page.locator('[aria-label*="Ratings and reviews" i]')
                    except Exception:
                        pass
                if container and container.count() > 0:
                    container.first.scroll_into_view_if_needed()
                    self._page.wait_for_timeout(700)
                    try:
                        text_parts.append(container.first.inner_text(timeout=1000))
                    except Exception:
                        pass
            except Exception:
                pass
            try:
                name_rx = re.compile(r"phone|tablet|chromebook|android\s*tv|tv|watch|wear(?:\s*os)?|car|android\s*auto|automotive|all devices", re.I)
                chips = self._page.get_by_role("button", name=name_rx)
                if chips.count() == 0:
                    try:
                        self._page.wait_for_selector('button', timeout=1000)
                        chips = self._page.get_by_role("button", name=name_rx)
                    except Exception:
                        pass
                chip_texts = []
                n = min(max(chips.count(), 0), 30)
                for i in range(n):
                    try:
                        t = chips.nth(i).inner_text(timeout=500)
                        if t and name_rx.search(t):
                            chip_texts.append(t)
                    except Exception:
                        continue
                if chip_texts:
                    text_parts.append(" | ".join(sorted(set(chip_texts), key=lambda x: x.lower())))
            except Exception:
                pass
            out_text = "\n".join([t for t in text_parts if t])
            try:
                with open(p, "w", encoding="utf-8") as f:
                    f.write(out_text)
            except Exception:
                pass
            time.sleep(self.delay)
            return out_text
        except Exception:
            return None


def read_xlsx_rows(path: str, sheet_name: Optional[str] = None) -> Tuple[List[Dict[str, str]], List[str]]:
    if not HAVE_OPENPYXL:
        raise RuntimeError("openpyxl is not installed. Please install openpyxl to read .xlsx input.")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []
    header = [str(h).strip() if h is not None else "" for h in header_row]
    seen = set()
    uniq_header: List[str] = []
    for i, h in enumerate(header):
        base = h or f"col_{i+1}"
        name = base
        k = 1
        while name.lower() in seen:
            k += 1
            name = f"{base}_{k}"
        seen.add(name.lower())
        uniq_header.append(name)
    out_rows: List[Dict[str, str]] = []
    for row in rows_iter:
        if row is None:
            continue
        d: Dict[str, str] = {}
        for i, h in enumerate(uniq_header):
            val = row[i] if i < len(row) else None
            if isinstance(val, bool):
                d[h] = "True" if val else ""
            elif val is None:
                d[h] = ""
            else:
                d[h] = str(val)
        out_rows.append(d)
    return out_rows, uniq_header


def resolve_column_map(header: List[str]) -> Dict[str, str]:
    lower_to_actual = {h.lower(): h for h in header}
    def pick(*cands: str, default: str) -> str:
        for c in cands:
            a = lower_to_actual.get(c.lower())
            if a:
                return a
        return default
    return {
        "mobile": pick("mobile", "phone", default=DEFAULT_COL_MAP["mobile"]),
        "tablet": pick("tablet", default=DEFAULT_COL_MAP["tablet"]),
        "chromebook": pick("chromebook", default=DEFAULT_COL_MAP["chromebook"]),
    "tv": pick("tv", "android tv", default=DEFAULT_COL_MAP["tv"]),
    "watch": pick("watch", "wear", "wear os", "wearos", default=DEFAULT_COL_MAP["watch"]),
    "car": pick("car", "android auto", "auto", "automotive", default=DEFAULT_COL_MAP["car"]),
    }


def detect_url_column(header: List[str], rows: List[Dict[str, str]]) -> Optional[str]:
    lowered = [h.lower() for h in header]
    candidates = []
    for h, low in zip(header, lowered):
        if any(k in low for k in ["google", "play", "url", "link"]) or low in ("url", "link"):
            candidates.append(h)
    for c in candidates + header:
        if not c:
            continue
        for r in rows[:200]:
            v = (r.get(c) or "").strip()
            if v.startswith(PLAY_PREFIX):
                return c
    for r in rows[:200]:
        for h in header:
            v = (r.get(h) or "").strip()
            if v.startswith(PLAY_PREFIX):
                return h
    return None


def extract_ratings_reviews_text(html: str) -> str:
    soup = None
    if BeautifulSoup:
        try:
            soup = BeautifulSoup(html, "lxml")
        except Exception:
            try:
                soup = BeautifulSoup(html, "html.parser")
            except Exception:
                soup = None
    if not soup:
        return ""

    # Known label variations
    labels = [
        "ratings and reviews",
        "ratings & reviews",
    ]
    def _norm(s: str) -> str:
        return re.sub(r"\s+", " ", s or "").strip().lower()

    label_rx = re.compile("|".join(re.escape(x) for x in labels), re.I)
    # aria-label match
    sec = soup.find(attrs={"aria-label": lambda v: isinstance(v, str) and label_rx.search(v)})
    if sec:
        return sec.get_text("\n", strip=True)

    # Heading match
    heading = soup.find(lambda t: t.name in ("h2", "h3") and label_rx.search(_norm(t.get_text(strip=True))))
    if heading:
        parts = [heading.get_text(" ", strip=True)]
        sib = heading.find_next_sibling()
        limit = 8
        while sib is not None and limit > 0:
            parts.append(sib.get_text("\n", strip=True))
            sib = sib.find_next_sibling()
            limit -= 1
        return "\n".join([p for p in parts if p])

    return ""


def is_trivial_section(text: str) -> bool:
    t = (text or "").strip()
    if not t:
        return True
    # Treat icon-only placeholders as trivial
    if re.fullmatch(r"arrow_forward", t.lower()):
        return True
    # Very short content is likely placeholder
    return len(t) < 20


def compile_token_regexes() -> Dict[str, re.Pattern]:
    def make_rx(words: List[str]) -> re.Pattern:
        # allow whitespace, NBSP, or hyphen between words, also zero-separator
        alts = []
        for w in words:
            parts = re.split(r"\s+", w.strip())
            sep = r"[\s\u00A0-]*"
            pat = sep.join(re.escape(p) for p in parts if p)
            alts.append(pat)
        return re.compile(rf"(?i)(?<![\w-])(?:{'|'.join(alts)})(?![\w-])")
    return {k: make_rx(v) for k, v in COMPAT_TOKENS.items()}


TOKEN_REGEXES = compile_token_regexes()


def detect_compat(info_text: str) -> Dict[str, bool]:
    text = (
        info_text
        .replace("\xa0", " ")
        .replace("\u202f", " ")
        .replace("\u2007", " ")
        .replace("\u2009", " ")
        .replace("\u200a", " ")
        .replace("\u00C2", "")
    )
    text = re.sub(r"\s+", " ", text).lower()
    found = {k: False for k in COMPAT_TOKENS.keys()}
    # Search only within the Ratings and reviews section text
    for key, rx in TOKEN_REGEXES.items():
        if rx.search(text):
            found[key] = True
    return found


def update_xlsx_inplace(src_path: str, dst_path: str, sheet_name: Optional[str], target_cols: List[str], rows: List[Dict[str, str]], blank_nontrue: bool = False) -> Tuple[int, List[str], str]:
    if not HAVE_OPENPYXL:
        raise RuntimeError("openpyxl is not installed. Please install openpyxl to update .xlsx in place.")
    wb = openpyxl.load_workbook(src_path, data_only=False)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    header_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=False))
    if not header_cells:
        wb.save(dst_path)
        return 0, [], dst_path
    header_values = [(str(c.value).strip() if c.value is not None else "") for c in header_cells[0]]
    lower_to_idx = {}
    for idx, name in enumerate(header_values, start=1):
        if name and name.lower() not in lower_to_idx:
            lower_to_idx[name.lower()] = idx
    present: List[Tuple[str, int]] = []
    missing: List[str] = []
    for col in target_cols:
        if not col:
            continue
        i = lower_to_idx.get(col.lower())
        if i is None:
            missing.append(col)
        else:
            present.append((col, i))
    updated_cells = 0
    max_rows = ws.max_row
    for r_idx in range(2, max_rows + 1):
        list_idx = r_idx - 2
        if list_idx >= len(rows):
            break
        row_dict = rows[list_idx]
        for col_name, c_idx in present:
            val = (row_dict.get(col_name) or "")
            is_true = str(val).strip().lower() == "true"
            cell = ws.cell(row=r_idx, column=c_idx)
            if is_true:
                if cell.value != "True":
                    cell.value = "True"
                    updated_cells += 1
            else:
                if blank_nontrue and cell.value not in (None, ""):
                    cell.value = None
                    updated_cells += 1
    try:
        wb.save(dst_path)
        effective_path = dst_path
    except PermissionError:
        base, ext = os.path.splitext(dst_path)
        alt = f"{base}-{int(time.time())}{ext}"
        wb.save(alt)
        effective_path = alt
    return updated_cells, missing, effective_path


def read_csv_rows(path: str) -> Tuple[List[Dict[str, str]], List[str], str]:
    with open(path, "r", encoding="utf-8", errors="ignore", newline="") as f:
        sample = f.read(4096)
        semi = sample.count(";")
        comma = sample.count(",")
        delimiter = ";" if semi >= comma else ","
        f.seek(0)
        reader = csv.DictReader(f, delimiter=delimiter)
        rows = list(reader)
        header = reader.fieldnames or []
    return rows, header, delimiter


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Enrich XLSX/CSV with Google Play device compatibility (Mobile/Tablet/Chromebook/TV/Watch/Car) from Ratings and reviews section.")
    parser.add_argument("input_csv", nargs="?", default="appsCombinedXLSX.xlsx", help="Path to input XLSX or CSV (default: appsCombinedXLSX.xlsx)")
    parser.add_argument("output_csv", nargs="?", default=None, help="Path to output CSV (default: <input>_android_compat.csv for CSV inputs)")
    parser.add_argument("--url-column", dest="url_col", default=None, help="Name of the column containing Google Play URLs (auto-detect if omitted)")
    parser.add_argument("--limit", type=int, default=0, help="Process only the first N Google Play URLs (for testing)")
    parser.add_argument("--resume", action="store_true", help="Skip rows where target columns already contain a True value")
    parser.add_argument("--delay", type=float, default=0.6, help="Delay in seconds between requests (default 0.6)")
    parser.add_argument("--cache-dir", default=os.path.join("cache", "google_pages"), help="Directory to cache fetched pages")
    parser.add_argument("--debug-url", dest="debug_url", default=None, help="Fetch a single Google Play URL and print the Ratings and reviews section text; skips table processing")
    parser.add_argument("--debug-detect", dest="debug_detect", default=None, help="Fetch a single Google Play URL and print detected flags and mapped columns; skips table processing")
    parser.add_argument("--render", action="store_true", help="Use Playwright to render pages when static HTML lacks Ratings and reviews content (requires 'pip install playwright' and 'playwright install')")
    parser.add_argument("--input-sheet", dest="input_sheet", default=None, help="XLSX sheet name to read (default: first sheet)")
    parser.add_argument("--blank-nontrue", action="store_true", help="When updating XLSX in-place, blank cells that are not True (default: leave as-is to preserve styles)")
    parser.add_argument("--output-xlsx", dest="output_xlsx", default=None, help="Also write results to this Excel .xlsx file for CSV inputs")
    parser.add_argument("--autosave-rows", dest="autosave_rows", type=int, default=0, help="If >0, write a progress CSV snapshot every N processed rows")

    args = parser.parse_args(argv)

    ensure_deps()

    # Debug single URL paths
    if args.debug_url:
        if not args.debug_url.startswith(PLAY_PREFIX):
            sys.stderr.write("--debug-url must be a Google Play URL (https://play.google.com...).\n")
            return 2
        s = make_session()
        html = load_page(args.debug_url, s, args.cache_dir, delay=args.delay)
        if not html:
            sys.stderr.write("Failed to load page.\n")
            return 1
        text = extract_ratings_reviews_text(html)
        if args.render:
            if not HAVE_PLAYWRIGHT:
                sys.stderr.write("Playwright not installed. Install with 'pip install playwright' then run 'playwright install'.\n")
            else:
                rt = load_page_rendered(args.debug_url, args.cache_dir, delay=args.delay)
                if rt and (is_trivial_section(text) or len(rt) > len(text)):
                    text = rt
        print("=== Ratings and reviews section text ===")
        print(text)
        return 0

    if args.debug_detect:
        if not args.debug_detect.startswith(PLAY_PREFIX):
            sys.stderr.write("--debug-detect must be a Google Play URL (https://play.google.com...).\n")
            return 2
        s = make_session()
        html = load_page(args.debug_detect, s, args.cache_dir, delay=args.delay)
        if not html:
            sys.stderr.write("Failed to load page.\n")
            return 1
        text = extract_ratings_reviews_text(html)
        if args.render:
            if not HAVE_PLAYWRIGHT:
                sys.stderr.write("Playwright not installed. Install with 'pip install playwright' then run 'playwright install'.\n")
            else:
                rt = load_page_rendered(args.debug_detect, args.cache_dir, delay=args.delay)
                if rt and (is_trivial_section(text) or len(rt) > len(text)):
                    text = rt
        flags = detect_compat(text)
        try:
            rows_tmp, header_tmp, _ = read_csv_rows(args.input_csv)
        except Exception:
            header_tmp = []
        col_map_tmp = resolve_column_map(header_tmp)
        print("=== Detected flags ===")
        for k in sorted(flags.keys()):
            print(f"{k}: {flags[k]}")
        print("=== Column mapping ===")
        for k in sorted(col_map_tmp.keys()):
            print(f"{k} -> {col_map_tmp[k]}")
        return 0

    is_xlsx_input = args.input_csv.lower().endswith(".xlsx")
    if is_xlsx_input:
        rows, header = read_xlsx_rows(args.input_csv, sheet_name=args.input_sheet)
        delimiter = ","
    else:
        rows, header, delimiter = read_csv_rows(args.input_csv)

    col_map = resolve_column_map(header)
    for col in col_map.values():
        if col not in header:
            header.append(col)

    url_col = args.url_col or detect_url_column(header, rows)
    if not url_col:
        sys.stderr.write("Could not detect the Google Play URL column. Use --url-column to specify.\n")
        return 2

    s = make_session()
    processed = 0
    updated = 0
    renderer = None
    if args.render and HAVE_PLAYWRIGHT:
        renderer = _PersistentRenderer(cache_dir=args.cache_dir, delay=args.delay)
        try:
            renderer.start()
        except Exception:
            renderer = None

    def row_complete(r: Dict[str, str]) -> bool:
        for col in col_map.values():
            v = (r.get(col) or "").strip()
            if v == "":
                return False
        return True

    # Determine autosave path if enabled
    autosave_path = None
    if args.autosave_rows and args.autosave_rows > 0:
        base = re.sub(r"\.xlsx$", "", args.input_csv, flags=re.I)
        autosave_path = f"{base}_android_progress.csv"

    for r in rows:
        url = (r.get(url_col) or "").strip()
        if not url.startswith(PLAY_PREFIX):
            continue
        if args.resume and row_complete(r):
            continue
        processed += 1
        if args.limit and processed > args.limit:
            break
        html = load_page(url, s, args.cache_dir, delay=args.delay)
        if not html:
            continue
        text = extract_ratings_reviews_text(html)
        # First, detect from static HTML
        flags_static = detect_compat(text)
        flags = flags_static
        # Only render if needed (no tokens found or section trivial)
        if args.render and (not any(flags_static.values()) or is_trivial_section(text)):
            rt = renderer.fetch(url) if renderer else load_page_rendered(url, args.cache_dir, delay=args.delay)
            if rt and (is_trivial_section(text) or len(rt) > len(text)):
                text = rt
                flags = detect_compat(text)
        changed = False
        for token, is_true in flags.items():
            out_col = col_map[token]
            if is_true and (r.get(out_col) or "").strip().lower() != "true":
                r[out_col] = "True"
                changed = True
        if changed:
            updated += 1
        if processed % 500 == 0:
            print(f"Progress: processed={processed}, updated_rows={updated}")
            sys.stdout.flush()
        if autosave_path and args.autosave_rows and (processed % args.autosave_rows == 0):
            try:
                with open(autosave_path, "w", encoding="utf-8", newline="") as f:
                    writer = csv.DictWriter(f, fieldnames=header)
                    writer.writeheader()
                    for rr in rows:
                        writer.writerow({k: rr.get(k, "") for k in header})
                print(f"Autosaved progress CSV: {autosave_path} (processed={processed})")
                sys.stdout.flush()
            except Exception as e:
                print(f"Autosave failed: {e}")
                sys.stdout.flush()

    if renderer:
        try:
            renderer.stop()
        except Exception:
            pass

    if is_xlsx_input:
        xlsx_path = args.output_xlsx or re.sub(r"\.xlsx$", "_android_compat.xlsx", args.input_csv, flags=re.I)
        target_cols = [col_map[k] for k in ("mobile", "tablet", "chromebook", "tv", "watch", "car")]
        _, missing_cols, effective = update_xlsx_inplace(
            args.input_csv, xlsx_path, args.input_sheet, target_cols, rows, blank_nontrue=args.blank_nontrue
        )
        print(f"Done. Processed Google Play rows: {processed}, updated rows: {updated}. Output: {effective}")
        if missing_cols:
            print(f"Note: columns not found in sheet (skipped): {', '.join(missing_cols)}")
        if args.output_csv:
            # optional CSV duplicate
            with open(args.output_csv, "w", encoding="utf-8", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=header)
                writer.writeheader()
                for rr in rows:
                    writer.writerow({k: rr.get(k, "") for k in header})
            print(f"Also wrote CSV: {args.output_csv}")
    else:
        out_path = args.output_csv or re.sub(r"\.csv$", "_android_compat.csv", args.input_csv, flags=re.I)
        with open(out_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=header, delimiter=delimiter)
            writer.writeheader()
            for rr in rows:
                writer.writerow({k: rr.get(k, "") for k in header})
        print(f"Done. Processed Google Play rows: {processed}, updated rows: {updated}. Output: {out_path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
