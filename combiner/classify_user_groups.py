import os
import re
import csv
import sys
import json
import time
import argparse
import hashlib
from typing import Dict, List, Optional, Tuple

# Lightweight, dependency-aligned with existing enrichers
try:
    from bs4 import BeautifulSoup  # type: ignore
    HAVE_BS4 = True
except Exception:
    HAVE_BS4 = False
    BeautifulSoup = None  # type: ignore

try:
    import requests  # type: ignore
    from requests.adapters import HTTPAdapter  # type: ignore
    from urllib3.util.retry import Retry  # type: ignore
    HAVE_REQ = True
except Exception:
    HAVE_REQ = False

try:
    import openpyxl  # type: ignore
    HAVE_OPENPYXL = True
except Exception:
    HAVE_OPENPYXL = False

# Optional render for dynamic Google Play descriptions
try:
    from playwright.sync_api import sync_playwright  # type: ignore
    HAVE_PLAYWRIGHT = True
except Exception:
    HAVE_PLAYWRIGHT = False

# Optional OpenAI/Azure OpenAI client; loaded lazily
try:
    from openai import OpenAI  # type: ignore
    HAVE_OPENAI = True
except Exception:
    HAVE_OPENAI = False


APPLE_PREFIX = "https://apps.apple.com"
PLAY_PREFIX = "https://play.google.com"


def ensure_deps():
    missing = []
    if not HAVE_BS4:
        missing.append("beautifulsoup4")
    if not HAVE_REQ:
        missing.append("requests")
    if not HAVE_OPENPYXL:
        missing.append("openpyxl")
    if missing:
        sys.stderr.write("Missing packages: " + ", ".join(missing) + "\n")
        sys.exit(2)


def make_session() -> "requests.Session":
    s = requests.Session()
    retries = Retry(
        total=3,
        backoff_factor=1.0,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET"],
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retries)
    s.mount("http://", adapter)
    s.mount("https://", adapter)
    s.headers.update(
        {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0.0.0 Safari/537.36"
            ),
            "Accept-Language": "en-US,en;q=0.9",
        }
    )
    return s


def sha1(s: str) -> str:
    return hashlib.sha1(s.encode("utf-8")).hexdigest()


def load_page(url: str, session: "requests.Session", cache_dir: str, delay: float = 0.6) -> Optional[str]:
    os.makedirs(cache_dir, exist_ok=True)
    key = sha1(url)
    p = os.path.join(cache_dir, f"{key}.html")
    if os.path.exists(p):
        try:
            with open(p, "r", encoding="utf-8", errors="ignore") as f:
                return f.read()
        except Exception:
            pass
    try:
        resp = session.get(url, timeout=25)
        if resp.status_code == 200 and resp.text:
            html = resp.text
            try:
                with open(p, "w", encoding="utf-8") as f:
                    f.write(html)
            except Exception:
                pass
            time.sleep(delay)
            return html
        else:
            if resp.status_code == 429:
                time.sleep(5)
            return None
    except Exception:
        return None


def load_page_rendered(url: str, cache_dir: str, delay: float = 0.6, timeout_ms: int = 10000) -> Optional[str]:
    if not HAVE_PLAYWRIGHT:
        return None
    os.makedirs(cache_dir, exist_ok=True)
    key = sha1(url + "|render|about")
    p = os.path.join(cache_dir, f"{key}.txt")
    if os.path.exists(p):
        try:
            with open(p, "r", encoding="utf-8", errors="ignore") as f:
                return f.read()
        except Exception:
            pass
    try:
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True)
            ua = (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0.0.0 Safari/537.36"
            )
            context = browser.new_context(
                user_agent=ua,
                locale="en-US",
                extra_http_headers={"Accept-Language": "en-US,en;q=0.9"},
            )
            page = context.new_page()
            target_url = url
            if "hl=" not in target_url:
                sep = "&" if "?" in target_url else "?"
                target_url = f"{target_url}{sep}hl=en&gl=us"
            page.goto(target_url, wait_until="domcontentloaded", timeout=timeout_ms)
            # Scroll and try to open "About this app"
            for _ in range(6):
                page.mouse.wheel(0, 1200)
                page.wait_for_timeout(250)
            text = ""
            try:
                heading = page.get_by_role("heading", name=re.compile("about this app", re.I))
                if heading and heading.count() > 0:
                    heading.first.scroll_into_view_if_needed()
                    page.wait_for_timeout(400)
                    # Try expanding if a more/arrow exists
                    more_btn = page.get_by_role("button", name=re.compile("more|arrow", re.I))
                    if more_btn and more_btn.count() > 0:
                        try:
                            more_btn.first.click(timeout=800)
                        except Exception:
                            pass
                    cont = heading.first.locator("xpath=following::*[self::div or self::section][1]")
                    if cont and cont.count() > 0:
                        text = cont.first.inner_text(timeout=1500)
            except Exception:
                pass
            if not text:
                # Fallback to page meta description
                try:
                    text = page.locator('meta[name="description"]').first.get_attribute("content") or ""
                except Exception:
                    text = ""
            try:
                with open(p, "w", encoding="utf-8") as f:
                    f.write(text or "")
            except Exception:
                pass
            page.close(); context.close(); browser.close()
            time.sleep(delay)
            return text or None
    except Exception:
        return None


def extract_apple_description(html: str) -> str:
    if not BeautifulSoup:
        return ""
    try:
        soup = BeautifulSoup(html, "lxml")
    except Exception:
        soup = BeautifulSoup(html, "html.parser")
    # Prefer the Description section
    try:
        heading = soup.find(lambda t: t.name in ("h2", "h3") and re.search(r"^description$", t.get_text(strip=True), re.I))
        if heading:
            sec = heading.find_next(lambda x: x.name in ("section", "div"))
            if sec:
                txt = sec.get_text("\n", strip=True)
                if len(txt) > 60:
                    return txt
    except Exception:
        pass
    # Fallback to meta description
    meta = soup.find("meta", attrs={"name": "description"})
    if meta and meta.get("content"):
        return str(meta.get("content"))
    meta = soup.find("meta", attrs={"property": "og:description"})
    if meta and meta.get("content"):
        return str(meta.get("content"))
    # Last resort: visible text around app header
    try:
        article = soup.find("article") or soup
        return article.get_text("\n", strip=True)[:2000]
    except Exception:
        return ""


def extract_google_description(html: str) -> str:
    if not BeautifulSoup:
        return ""
    try:
        soup = BeautifulSoup(html, "lxml")
    except Exception:
        soup = BeautifulSoup(html, "html.parser")
    # Try About this app area
    label_rx = re.compile(r"about this app", re.I)
    try:
        heading = soup.find(lambda t: t.name in ("h2", "h3") and label_rx.search(t.get_text(" ", strip=True)))
        if heading:
            parts = [heading.get_text(" ", strip=True)]
            sib = heading.find_next_sibling()
            limit = 6
            while sib is not None and limit > 0:
                parts.append(sib.get_text("\n", strip=True))
                sib = sib.find_next_sibling()
                limit -= 1
            text = "\n".join([p for p in parts if p])
            if len(text) > 60:
                return text
    except Exception:
        pass
    # Fallback to meta description
    meta = soup.find("meta", attrs={"name": "description"})
    if meta and meta.get("content"):
        return str(meta.get("content"))
    meta = soup.find("meta", attrs={"property": "og:description"})
    if meta and meta.get("content"):
        return str(meta.get("content"))
    return ""


def norm_text(s: str) -> str:
    s = (s or "").replace("\xa0", " ").replace("\u202f", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def read_xlsx_rows(path: str, sheet_name: Optional[str] = None) -> Tuple[List[Dict[str, str]], List[str]]:
    if not HAVE_OPENPYXL:
        raise RuntimeError("openpyxl is not installed")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []
    header = [str(h).strip() if h is not None else "" for h in header_row]
    seen = set()
    uniq_header: List[str] = []
    for i, h in enumerate(header):
        base = h or f"col_{i+1}"
        name = base
        k = 1
        while name.lower() in seen:
            k += 1
            name = f"{base}_{k}"
        seen.add(name.lower())
        uniq_header.append(name)
    out_rows: List[Dict[str, str]] = []
    for row in rows_iter:
        if row is None:
            continue
        d: Dict[str, str] = {}
        for i, h in enumerate(uniq_header):
            val = row[i] if i < len(row) else None
            if isinstance(val, bool):
                d[h] = "True" if val else ""
            elif val is None:
                d[h] = ""
            else:
                d[h] = str(val)
        out_rows.append(d)
    return out_rows, uniq_header


def read_csv_rows(path: str) -> Tuple[List[Dict[str, str]], List[str], str]:
    with open(path, "r", encoding="utf-8", errors="ignore", newline="") as f:
        sample = f.read(4096)
        semi = sample.count(";")
        comma = sample.count(",")
        delimiter = ";" if semi >= comma else ","
        f.seek(0)
        reader = csv.DictReader(f, delimiter=delimiter)
        rows = list(reader)
        header = reader.fieldnames or []
    return rows, header, delimiter


def detect_url_column(header: List[str], rows: List[Dict[str, str]]) -> Optional[str]:
    lowered = [h.lower() for h in header]
    candidates = []
    for h, low in zip(header, lowered):
        if any(k in low for k in ["url", "link", "apple", "itunes", "google", "play"]):
            candidates.append(h)
    for c in candidates + header:
        if not c:
            continue
        for r in rows[:200]:
            v = (r.get(c) or "").strip()
            if v.startswith(APPLE_PREFIX) or v.startswith(PLAY_PREFIX):
                return c
    for r in rows[:200]:
        for h in header:
            v = (r.get(h) or "").strip()
            if v.startswith(APPLE_PREFIX) or v.startswith(PLAY_PREFIX):
                return h
    return None


def classify_with_openai(client: "OpenAI", model: str, text: str, labels: List[str], temperature: float = 0.0, max_retries: int = 2, rpm: int = 60) -> str:
    # Rate limit: simple sleep per call
    if rpm > 0:
        time.sleep(60.0 / float(rpm))
    sys_prompt = (
        "You are a precise classifier for mobile app audiences. Given an app description, "
        "return the single best-fitting label from the allowed set. If none fits, use 'Other'. "
        "Output strict JSON: {\"label\": \"<one label>\"}."
    )
    user_prompt = (
        "Allowed labels: " + ", ".join(labels) + "\n\n" +
        "Description:\n" + text.strip()[:8000]
    )
    last_err = None
    for _ in range(max_retries + 1):
        try:
            resp = client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": sys_prompt},
                    {"role": "user", "content": user_prompt},
                ],
                temperature=temperature,
                max_tokens=20,
            )
            content = resp.choices[0].message.content or ""
            # Try to parse JSON; fall back to raw string
            label = parse_label_from_content(content, labels)
            return label
        except Exception as e:
            last_err = e
            time.sleep(1.5)
    raise RuntimeError(f"OpenAI classification failed: {last_err}")


def get_openai_client(provider: str, model: str, azure_endpoint: Optional[str], azure_api_version: Optional[str]) -> "OpenAI":
    # Standard OpenAI
    if provider == "openai":
        if not HAVE_OPENAI:
            raise RuntimeError("openai package not installed. Install with 'pip install openai'.")
        api_key = os.environ.get("OPENAI_API_KEY")
        if not api_key:
            raise RuntimeError("Set OPENAI_API_KEY in environment.")
        return OpenAI(api_key=api_key)
    # Azure OpenAI via OpenAI SDK-compatible endpoint
    if provider == "azure":
        if not HAVE_OPENAI:
            raise RuntimeError("openai package not installed. Install with 'pip install openai'.")
        endpoint = (azure_endpoint or os.environ.get("AZURE_OPENAI_ENDPOINT") or "").rstrip("/")
        api_key = os.environ.get("AZURE_OPENAI_API_KEY")
        api_version = azure_api_version or os.environ.get("AZURE_OPENAI_API_VERSION") or "2024-05-01-preview"
        if not endpoint or not api_key:
            raise RuntimeError("Set AZURE_OPENAI_ENDPOINT and AZURE_OPENAI_API_KEY in environment (and optionally AZURE_OPENAI_API_VERSION).")
        # Expect 'model' to be the deployment name for Azure
        base_url = f"{endpoint}/openai/deployments/{model}"
        return OpenAI(api_key=api_key, base_url=base_url, default_headers={"api-key": api_key, "x-ms-version": api_version})
    raise RuntimeError("Unknown provider. Use 'openai' or 'azure'.")


def parse_label_from_content(content: str, labels: List[str]) -> str:
    # Try JSON
    try:
        j = json.loads(content)
        val = str(j.get("label", "")).strip()
        if val:
            return choose_label(val, labels)
    except Exception:
        pass
    # Extract bare word
    m = re.search(r"label\"?\s*[:=]\s*\"([^\"]+)\"", content, re.I)
    if m:
        return choose_label(m.group(1), labels)
    # Last resort: first matching allowed label substring
    for L in labels:
        if re.search(rf"\b{re.escape(L)}\b", content, re.I):
            return L
    return "Other" if "Other" in labels else labels[0]


def choose_label(raw: str, labels: List[str]) -> str:
    r = raw.strip()
    # Exact/ci match
    for L in labels:
        if r.lower() == L.lower():
            return L
    # Heuristic normalize (singular/plural, casing)
    r1 = re.sub(r"s$", "", r.lower())
    for L in labels:
        if r1 == re.sub(r"s$", "", L.lower()):
            return L
    return "Other" if "Other" in labels else labels[0]


def update_xlsx_text_inplace(src_path: str, dst_path: str, sheet_name: Optional[str], target_col: str, rows: List[Dict[str, str]]) -> Tuple[int, str]:
    if not HAVE_OPENPYXL:
        raise RuntimeError("openpyxl is not installed")
    wb = openpyxl.load_workbook(src_path, data_only=False)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    header_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=False))
    if not header_cells:
        wb.save(dst_path)
        return 0, dst_path
    header_values = [(str(c.value).strip() if c.value is not None else "") for c in header_cells[0]]
    lower_to_idx: Dict[str, int] = {}
    for idx, name in enumerate(header_values, start=1):
        if name and name.lower() not in lower_to_idx:
            lower_to_idx[name.lower()] = idx
    c_idx = lower_to_idx.get(target_col.lower())
    if c_idx is None:
        # Append column
        c_idx = len(header_values) + 1
        ws.cell(row=1, column=c_idx).value = target_col
    updated = 0
    max_rows = ws.max_row
    for r_idx in range(2, max_rows + 1):
        list_idx = r_idx - 2
        if list_idx >= len(rows):
            break
        val = rows[list_idx].get(target_col, "")
        cell = ws.cell(row=r_idx, column=c_idx)
        if (cell.value or "") != val:
            cell.value = val
            updated += 1
    try:
        wb.save(dst_path)
        effective = dst_path
    except PermissionError:
        base, ext = os.path.splitext(dst_path)
        alt = f"{base}-{int(time.time())}{ext}"
        wb.save(alt)
        effective = alt
    return updated, effective


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Classify main user group for apps using OpenAI/Azure from Apple/Google Play descriptions.")
    parser.add_argument("input", nargs="?", default="appsCombinedXLSX.xlsx", help="Input XLSX or CSV (default: appsCombinedXLSX.xlsx)")
    parser.add_argument("output", nargs="?", default=None, help="Optional output CSV path for CSV inputs; for XLSX, a _classified.xlsx is written")
    parser.add_argument("--input-sheet", dest="input_sheet", default=None, help="Sheet name for XLSX inputs")
    parser.add_argument("--url-column", dest="url_col", default=None, help="Column name containing store URLs (auto-detect if omitted)")
    parser.add_argument("--label-column", dest="label_col", default="MainUserGroup", help="Output column name for the predicted label (default: MainUserGroup)")
    parser.add_argument("--labels", dest="labels", default="Athlete,Support staff,Supporter,Coach,Official,General,Other", help="Comma-separated allowed labels (default: a common set incl. Other)")
    parser.add_argument("--limit", type=int, default=0, help="Process only first N URLs (testing)")
    parser.add_argument("--resume", action="store_true", help="Skip rows where label column already has a value")
    parser.add_argument("--delay", type=float, default=0.6, help="Delay between HTTP requests (default 0.6)")
    parser.add_argument("--cache-dir", default=os.path.join("cache", "descriptions"), help="Directory to cache fetched pages and rendered text")
    parser.add_argument("--render", action="store_true", help="Use Playwright to render Google Play when static HTML lacks description")
    # LLM settings
    parser.add_argument("--provider", choices=["openai", "azure"], default="openai", help="LLM provider (openai or azure)")
    parser.add_argument("--model", default="gpt-4o-mini", help="Model (OpenAI model name or Azure deployment name)")
    parser.add_argument("--temperature", type=float, default=0.0, help="Sampling temperature (default 0.0)")
    parser.add_argument("--rpm", type=int, default=60, help="API calls per minute throttle (default 60)")
    parser.add_argument("--azure-endpoint", dest="azure_endpoint", default=None, help="Azure OpenAI endpoint (if provider=azure)")
    parser.add_argument("--azure-api-version", dest="azure_api_version", default=None, help="Azure OpenAI API version (optional)")
    parser.add_argument("--debug-url", dest="debug_url", default=None, help="Fetch one URL, print extracted description and predicted label")

    args = parser.parse_args(argv)

    ensure_deps()

    labels = [x.strip() for x in args.labels.split(",") if x.strip()]
    if "Other" not in labels:
        labels.append("Other")

    # Initialize LLM client
    client = get_openai_client(args.provider, args.model, args.azure_endpoint, args.azure_api_version)

    s = make_session()

    def fetch_description(url: str) -> str:
        html = load_page(url, s, args.cache_dir, delay=args.delay)
        if not html:
            return ""
        if url.startswith(APPLE_PREFIX):
            return norm_text(extract_apple_description(html))
        if url.startswith(PLAY_PREFIX):
            text = norm_text(extract_google_description(html))
            if args.render and len(text) < 60:
                rt = load_page_rendered(url, args.cache_dir, delay=args.delay)
                if rt and len(rt) > len(text):
                    text = norm_text(rt)
            return text
        return ""

    # Single-URL debug flow
    if args.debug_url:
        url = args.debug_url.strip()
        if not (url.startswith(APPLE_PREFIX) or url.startswith(PLAY_PREFIX)):
            sys.stderr.write("--debug-url must be an Apple or Google Play URL.\n")
            return 2
        desc = fetch_description(url)
        if not desc:
            print("No description extracted.")
            return 1
        label = classify_with_openai(client, args.model, desc, labels, temperature=args.temperature, rpm=args.rpm)
        print("=== Extracted description (truncated) ===")
        print(desc[:1200])
        print("\n=== Predicted label ===")
        print(label)
        return 0

    # Bulk flow
    is_xlsx_input = args.input.lower().endswith(".xlsx")
    if is_xlsx_input:
        rows, header = read_xlsx_rows(args.input, sheet_name=args.input_sheet)
        delimiter = ","
    else:
        rows, header, delimiter = read_csv_rows(args.input)

    # Ensure label column present in header ordering
    if args.label_col not in header:
        header.append(args.label_col)

    url_col = args.url_col or detect_url_column(header, rows)
    if not url_col:
        sys.stderr.write("Could not detect URL column. Use --url-column.\n")
        return 2

    # Classification cache (by description sha1)
    cache_json_path = os.path.join(args.cache_dir, "classify_cache.jsonl")
    os.makedirs(args.cache_dir, exist_ok=True)
    classify_cache: Dict[str, str] = {}
    if os.path.exists(cache_json_path):
        try:
            with open(cache_json_path, "r", encoding="utf-8", errors="ignore") as f:
                for line in f:
                    try:
                        obj = json.loads(line)
                        if isinstance(obj, dict) and "key" in obj and "label" in obj:
                            classify_cache[str(obj["key"])] = str(obj["label"])
                    except Exception:
                        continue
        except Exception:
            pass

    def cache_get(desc: str) -> Optional[str]:
        k = sha1(desc)
        return classify_cache.get(k)

    def cache_put(desc: str, label: str):
        k = sha1(desc)
        classify_cache[k] = label
        try:
            with open(cache_json_path, "a", encoding="utf-8") as f:
                f.write(json.dumps({"key": k, "label": label}, ensure_ascii=False) + "\n")
        except Exception:
            pass

    processed = 0
    updated = 0

    def row_has_label(r: Dict[str, str]) -> bool:
        return bool((r.get(args.label_col) or "").strip())

    for r in rows:
        url = (r.get(url_col) or "").strip()
        if not (url.startswith(APPLE_PREFIX) or url.startswith(PLAY_PREFIX)):
            continue
        if args.resume and row_has_label(r):
            continue
        processed += 1
        if args.limit and processed > args.limit:
            break
        desc = fetch_description(url)
        if not desc:
            continue
        cached = cache_get(desc)
        if cached:
            label = cached
        else:
            label = classify_with_openai(client, args.model, desc, labels, temperature=args.temperature, rpm=args.rpm)
            cache_put(desc, label)
        if (r.get(args.label_col) or "") != label:
            r[args.label_col] = label
            updated += 1

    if is_xlsx_input:
        out_xlsx = re.sub(r"\.xlsx$", "_classified.xlsx", args.input, flags=re.I)
        # Reuse Excel update to preserve formatting: write only the label column
        # Simple implementation: open workbook and update one column by header name
        # Build rows aligned with existing order
        u, effective = update_xlsx_text_inplace(args.input, out_xlsx, args.input_sheet, args.label_col, rows)
        print(f"Done. Processed: {processed}, updated: {updated}. Output: {effective}")
    else:
        out_csv = args.output or re.sub(r"\.csv$", "_classified.csv", args.input, flags=re.I)
        with open(out_csv, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=header, delimiter=delimiter)
            writer.writeheader()
            for rr in rows:
                writer.writerow({k: rr.get(k, "") for k in header})
        print(f"Done. Processed: {processed}, updated: {updated}. Output: {out_csv}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
