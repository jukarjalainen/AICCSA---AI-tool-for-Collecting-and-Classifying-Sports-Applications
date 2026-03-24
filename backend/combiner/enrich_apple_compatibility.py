import csv
import os
import re
import sys
import time
import hashlib
import argparse
from typing import Dict, List, Optional, Tuple
from urllib.parse import urlparse, parse_qsl, urlencode, urlunparse

try:
    # Prefer lxml if available for robustness; fallback to html.parser
    from bs4 import BeautifulSoup
    HAVE_BS4 = True
except Exception:  # pragma: no cover
    HAVE_BS4 = False
    BeautifulSoup = None  # type: ignore

try:
    import requests
    from requests.adapters import HTTPAdapter
    from urllib3.util.retry import Retry
    HAVE_REQ = True
except Exception:  # pragma: no cover
    HAVE_REQ = False

# Optional dependency for XLSX export
try:
    import openpyxl
    from openpyxl.worksheet.table import Table, TableStyleInfo
    HAVE_OPENPYXL = True
except Exception:  # pragma: no cover
    HAVE_OPENPYXL = False


APP_STORE_PREFIX = "https://apps.apple.com"

# Map canonical compatibility tokens -> expected CSV column names from the user's sheet
# Default mapping; we'll resolve certain columns (like iPod/iPad touch) against the actual header later
COMPAT_COLUMN_MAP = {
    "ipad": "Ipad",
    "ipod touch": "Ipod touch",
    "iphone": "iPhone",
    "mac": "Mac",
    "apple tv": "Apple TV",
    "apple watch": "Apple watch",
    "apple vision pro": "Apple Vision",
}

# Precompiled regexes to detect tokens as words (case-insensitive)
def _token_to_pattern(token: str) -> re.Pattern:
    # Allow space, NBSP, or hyphen between words to catch variants like "iPod touch" or "iPod-touch"
    parts = re.split(r"\s+", token.strip())
    # Use * to also allow zero separator in rare cases like "ipodtouch"
    sep = r"[\s\u00A0-]*"
    pat = sep.join(re.escape(p) for p in parts if p)
    return re.compile(rf"(?i)(?<![\w-]){pat}(?![\w-])")


TOKEN_REGEXES = {token: _token_to_pattern(token) for token in COMPAT_COLUMN_MAP.keys()}


def resolve_column_map(header: List[str]) -> Dict[str, str]:
    """Resolve target column names against the CSV header (case-insensitive).
    Prefer existing header spellings; for iPod/iPad touch, accept either variant.
    """
    lower_to_actual = {h.lower(): h for h in header}
    col_map: Dict[str, str] = {}

    def pick_actual(*candidates: str) -> Optional[str]:
        for cand in candidates:
            actual = lower_to_actual.get(cand.lower())
            if actual:
                return actual
        return None

    # iPad column (cover iPad vs Ipad)
    col_map["ipad"] = pick_actual("iPad", "Ipad") or COMPAT_COLUMN_MAP["ipad"]
    # iPod/iPad touch column (support multiple header spellings + regex fallback for odd spacing)
    ipod_col = pick_actual("iPod touch", "Ipod touch", "iPad touch", "Ipad touch")
    if not ipod_col:
        # Regex fallback: find any header that looks like "ip?d touch" allowing extra spaces
        for h in header:
            if re.search(r"(?i)\bip[oa]d\s*touch\b", h.strip()):
                ipod_col = h
                break
    col_map["ipod touch"] = ipod_col or COMPAT_COLUMN_MAP["ipod touch"]
    # iPhone column (no fallback to Mobile)
    col_map["iphone"] = pick_actual("iPhone") or COMPAT_COLUMN_MAP["iphone"]
    # Other straightforward columns
    col_map["mac"] = pick_actual("Mac") or COMPAT_COLUMN_MAP["mac"]
    col_map["apple tv"] = pick_actual("Apple TV", "Apple tv") or COMPAT_COLUMN_MAP["apple tv"]
    col_map["apple watch"] = pick_actual("Apple Watch", "Apple watch") or COMPAT_COLUMN_MAP["apple watch"]
    col_map["apple vision pro"] = pick_actual("Apple Vision", "Apple Vision Pro") or COMPAT_COLUMN_MAP["apple vision pro"]

    return col_map


def make_session() -> requests.Session:
    s = requests.Session()
    retries = Retry(
        total=3,
        backoff_factor=1.2,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET"],
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retries)
    s.mount("http://", adapter)
    s.mount("https://", adapter)
    s.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/125.0.0.0 Safari/537.36"
        ),
        "Accept-Language": "en-US,en;q=0.9",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
    })
    return s


def ensure_bs4():
    if not HAVE_BS4:
        sys.stderr.write("BeautifulSoup4 not installed. Please install beautifulsoup4.\n")
        sys.exit(2)
    if not HAVE_REQ:
        sys.stderr.write("requests not installed. Please install requests.\n")
        sys.exit(2)


def read_csv_rows(path: str) -> Tuple[List[Dict[str, str]], List[str], str]:
    # Auto-detect delimiter between semicolon and comma
    with open(path, "r", encoding="utf-8", errors="ignore", newline="") as f:
        sample = f.read(4096)
        semi = sample.count(";")
        comma = sample.count(",")
        delimiter = ";" if semi >= comma else ","
        f.seek(0)
        reader = csv.DictReader(f, delimiter=delimiter)
        rows = list(reader)
        header = reader.fieldnames or []
    return rows, header, delimiter


def read_xlsx_rows(path: str, sheet_name: Optional[str] = None) -> Tuple[List[Dict[str, str]], List[str]]:
    if not HAVE_OPENPYXL:
        raise RuntimeError("openpyxl is not installed. Please install openpyxl to read .xlsx input.")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []
    header = [str(h).strip() if h is not None else "" for h in header_row]
    # Ensure unique header names
    seen = set()
    uniq_header: List[str] = []
    for i, h in enumerate(header):
        base = h or f"col_{i+1}"
        name = base
        k = 1
        while name.lower() in seen:
            k += 1
            name = f"{base}_{k}"
        seen.add(name.lower())
        uniq_header.append(name)
    out_rows: List[Dict[str, str]] = []
    for row in rows_iter:
        if row is None:
            continue
        d: Dict[str, str] = {}
        for i, h in enumerate(uniq_header):
            val = row[i] if i < len(row) else None
            if isinstance(val, bool):
                d[h] = "True" if val else ""
            elif val is None:
                d[h] = ""
            else:
                d[h] = str(val)
        out_rows.append(d)
    return out_rows, uniq_header


def write_csv_rows(path: str, header: List[str], rows: List[Dict[str, str]], delimiter: str) -> None:
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=header, delimiter=delimiter, quoting=csv.QUOTE_MINIMAL)
        writer.writeheader()
        for r in rows:
            writer.writerow({k: r.get(k, "") for k in header})


def write_xlsx_rows(path: str, header: List[str], rows: List[Dict[str, str]], boolean_cols: List[str]) -> None:
    if not HAVE_OPENPYXL:
        raise RuntimeError("openpyxl is not installed. Please install openpyxl to use --output-xlsx.")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Apps"
    # Write header
    for c, h in enumerate(header, start=1):
        ws.cell(row=1, column=c, value=h)
    # Write rows with boolean coercion
    bool_set = {b.lower() for b in boolean_cols}
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, h in enumerate(header, start=1):
            val = row.get(h, "")
            if h.lower() in bool_set:
                v = str(val).strip().lower() == "true"
                ws.cell(row=r_idx, column=c_idx, value=v)
            else:
                ws.cell(row=r_idx, column=c_idx, value=val)
    # Add Excel table for nicer filters
    last_col_letter = openpyxl.utils.get_column_letter(len(header))
    last_row = len(rows) + 1
    table_ref = f"A1:{last_col_letter}{last_row}"
    tbl = Table(displayName="AppsTable", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
                           showRowStripes=True, showColumnStripes=False)
    tbl.tableStyleInfo = style
    ws.add_table(tbl)
    wb.save(path)


def update_xlsx_inplace(src_path: str, dst_path: str, sheet_name: Optional[str], target_cols: List[str], rows: List[Dict[str, str]], blank_nontrue: bool = False) -> Tuple[int, List[str], str]:
    """Update an existing XLSX workbook in-place style (preserving validations/formatting).

    - Only writes to the provided target_cols headers.
    - Writes "True" (string) when the corresponding value in rows is True (case-insensitive),
      otherwise leaves the existing cell value unchanged (so it can remain empty).
    - Saves to dst_path (can be different from src_path) without altering sheet structure.

    Returns (updated_cells_count, missing_columns, effective_output_path) for quick diagnostics.
    """
    if not HAVE_OPENPYXL:
        raise RuntimeError("openpyxl is not installed. Please install openpyxl to update .xlsx in place.")

    wb = openpyxl.load_workbook(src_path, data_only=False)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    # Read header from first row
    header_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=False))
    if not header_cells:
        wb.save(dst_path)
        return 0, [], dst_path
    header_values = [
        (str(c.value).strip() if c.value is not None else "")
        for c in header_cells[0]
    ]
    # Map case-insensitive header -> 1-based column index
    lower_to_idx = {}
    for idx, name in enumerate(header_values, start=1):
        if name and name.lower() not in lower_to_idx:
            lower_to_idx[name.lower()] = idx

    # Resolve which target columns exist in the sheet
    present: List[Tuple[str, int]] = []
    missing: List[str] = []
    for col in target_cols:
        if not col:
            continue
        idx = lower_to_idx.get(col.lower())
        if idx is None:
            missing.append(col)
        else:
            present.append((col, idx))

    updated_cells = 0
    max_rows = ws.max_row
    # Iterate worksheet rows alongside our rows data; assume same order/length
    for r_idx in range(2, max_rows + 1):
        list_idx = r_idx - 2
        if list_idx >= len(rows):
            break
        row_dict = rows[list_idx]
        for col_name, c_idx in present:
            val = (row_dict.get(col_name) or "")
            is_true = str(val).strip().lower() == "true"
            cell = ws.cell(row=r_idx, column=c_idx)
            if is_true:
                # Write text "True" to play nicely with dropdown list validations expecting text
                if cell.value != "True":
                    cell.value = "True"
                    updated_cells += 1
            else:
                # Not true: default is to leave the cell as-is to preserve any styling.
                # Optionally blank it if requested.
                if blank_nontrue and cell.value not in (None, ""):
                    cell.value = None
                    updated_cells += 1

    # Save with fallback when the destination is locked (e.g., open in Excel)
    try:
        wb.save(dst_path)
        effective_path = dst_path
    except PermissionError:
        base, ext = os.path.splitext(dst_path)
        alt = f"{base}-{int(time.time())}{ext}"
        wb.save(alt)
        effective_path = alt
    return updated_cells, missing, effective_path


def detect_url_column(header: List[str], rows: List[Dict[str, str]]) -> Optional[str]:
    # Try heuristics on header names
    lowered = [h.lower() for h in header]
    candidates = []
    for h, low in zip(header, lowered):
        if any(k in low for k in ["appstore", "app_store", "apple", "url"]) or low in ("url", "link"):
            candidates.append(h)
    # Validate candidates by checking actual values
    for c in candidates + header:
        if not c:
            continue
        for r in rows[:200]:
            v = (r.get(c) or "").strip()
            if v.startswith(APP_STORE_PREFIX):
                return c
    # Last resort: scan any column for a value starting with apps.apple.com
    for r in rows[:200]:
        for h in header:
            v = (r.get(h) or "").strip()
            if v.startswith(APP_STORE_PREFIX):
                return h
    return None


def sha1(s: str) -> str:
    return hashlib.sha1(s.encode("utf-8")).hexdigest()


def load_page(url: str, session: requests.Session, cache_dir: str, delay: float = 0.6) -> Optional[str]:
    os.makedirs(cache_dir, exist_ok=True)
    key = sha1(url)
    p = os.path.join(cache_dir, f"{key}.html")
    if os.path.exists(p):
        try:
            with open(p, "r", encoding="utf-8", errors="ignore") as f:
                return f.read()
        except Exception:
            pass
    try:
        resp = session.get(url, timeout=20)
        if resp.status_code == 200 and resp.text:
            html = resp.text
            try:
                with open(p, "w", encoding="utf-8") as f:
                    f.write(html)
            except Exception:
                pass
            time.sleep(delay)
            return html
        else:
            # Backoff on 429
            if resp.status_code == 429:
                time.sleep(5)
            return None
    except Exception:
        return None


def with_platform(url: str, platform: str) -> str:
    try:
        p = urlparse(url)
        q = dict(parse_qsl(p.query, keep_blank_values=True))
        if q.get("platform") == platform:
            return url
        q["platform"] = platform
        new_q = urlencode(q)
        return urlunparse((p.scheme, p.netloc, p.path, p.params, new_q, p.fragment))
    except Exception:
        return url


def extract_information_text(html: str) -> str:
    soup = None
    if BeautifulSoup:
        try:
            soup = BeautifulSoup(html, "lxml")
        except Exception:
            try:
                soup = BeautifulSoup(html, "html.parser")
            except Exception:
                soup = None
    if not soup:
        # Can't parse HTML; to avoid false positives, return empty so no flags are set
        return ""

    # Known localized labels for the Information section (common locales)
    info_labels = [
        "information",        # en
        "informations",       # fr
        "informazioni",       # it
        "información",        # es
        "informacoes",        # pt (no diacritics fallback)
        "informações",        # pt-BR/pt-PT
        "informationen",      # de
        "informatie",         # nl
        "情報",                # ja
        "信息",                # zh-CN
        "資訊",                # zh-TW/HK
        "정보",                # ko
        "информация",         # ru
    ]

    def _norm(s: str) -> str:
        return re.sub(r"\s+", " ", s or "").strip().lower()

    label_rx = re.compile("|".join(re.escape(lbl) for lbl in info_labels), re.I)
    label_prefix_rx = re.compile(r"^informa", re.I)

    # Strategy 1: section with aria-label matching localized labels or starting with "Informa"
    info = soup.find(attrs={
        "aria-label": lambda v: isinstance(v, str) and (label_rx.search(v) or label_prefix_rx.search(v))
    })
    if info:
        return info.get_text("\n", strip=True)

    # Strategy 2: heading text equals one of the localized labels, then sibling content
    known_set = {_norm(l) for l in info_labels}
    heading = soup.find(
        lambda t: t.name in ("h2", "h3") and (
            _norm(t.get_text(strip=True)) in known_set or _norm(t.get_text(strip=True)).startswith("informa")
        )
    )
    if heading:
        # take the parent/next siblings
        parts = [heading.get_text(" ", strip=True)]
        sib = heading.find_next_sibling()
        limit = 10
        while sib is not None and limit > 0:
            parts.append(sib.get_text("\n", strip=True))
            sib = sib.find_next_sibling()
            limit -= 1
        return "\n".join([p for p in parts if p])

    # Strategy 3: do not fallback to whole page; if we can't find Information, treat as unknown
    return ""


def detect_compat(info_text: str) -> Dict[str, bool]:
    # Normalize NBSP and whitespace, then lowercase
    # Normalize common Unicode spaces to ASCII space
    text = (
        info_text
        .replace("\xa0", " ")  # NBSP
        .replace("\u202f", " ")  # NARROW NBSP
        .replace("\u2007", " ")  # FIGURE SPACE
        .replace("\u2009", " ")  # THIN SPACE
        .replace("\u200a", " ")  # HAIR SPACE
        .replace("\u00C2", "")   # Strip stray 'Â' from mojibake of NBSP
    )
    text = re.sub(r"\s+", " ", text)
    text = text.lower()
    # Track by canonical tokens; map to output columns later
    found: Dict[str, bool] = {t: False for t in COMPAT_COLUMN_MAP}

    # Prioritize searching near a "compatibility" keyword if present
    compat_region = text
    m = re.search(r"compatib(?:le|ility)", text)
    if m:
        start = max(0, m.start() - 4000)
        end = min(len(text), m.end() + 8000)
        compat_region = text[start:end]

    # Normalize whitespace inside the focused region to catch split words like "ipod\n touch"
    compat_region = re.sub(r"\s+", " ", compat_region)

    for token, rx in TOKEN_REGEXES.items():
        if rx.search(compat_region):
            found[token] = True

    # No full-page fallback scanning: only the Information section determines compatibility

    # Heuristics: sometimes Mac is referenced via "macOS"; Apple TV via "tvOS"; Apple Watch via "watchOS"; Vision via "visionOS"
    if not found["mac"] and re.search(r"\bmacos\b", compat_region):
        found["mac"] = True
    if not found["apple tv"] and re.search(r"\btvos\b", compat_region):
        found["apple tv"] = True
    if not found["apple watch"] and re.search(r"\bwatchos\b", compat_region):
        found["apple watch"] = True
    if not found["apple vision pro"] and re.search(r"\bvisionos\b|\bvision pro\b", compat_region):
        found["apple vision pro"] = True

    # iPhone presence (only within Information section) implies Mobile=True
    if not found.get("iphone", False):
        if re.search(r"(?i)(?<![\w-])iphone(?![\w-])", compat_region):
            found["iphone"] = True

    return found


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Enrich CSV/XLSX with Apple App Store compatibility flags by scraping the Information section.")
    parser.add_argument("input_csv", nargs="?", default="appsCombinedCSV.csv", help="Path to input CSV or XLSX (default: appsCombinedCSV.csv)")
    parser.add_argument("output_csv", nargs="?", default=None, help="Path to output CSV (default: <input> with _compat.csv suffix for CSV inputs)")
    parser.add_argument("--url-column", dest="url_col", default=None, help="Name of the column containing App Store URLs (auto-detect if omitted)")
    parser.add_argument("--limit", type=int, default=0, help="Process only the first N App Store URLs (for testing)")
    parser.add_argument("--resume", action="store_true", help="Skip rows where target columns already contain a True value")
    parser.add_argument("--delay", type=float, default=0.6, help="Delay in seconds between requests (default 0.6)")
    parser.add_argument("--cache-dir", default=os.path.join("cache", "apple_pages"), help="Directory to cache fetched pages")
    parser.add_argument("--debug-url", dest="debug_url", default=None, help="Fetch a single App Store URL and print the Information section text; skips CSV processing")
    parser.add_argument("--debug-detect", dest="debug_detect", default=None, help="Fetch a single App Store URL and print detected flags and mapped columns; skips CSV processing")
    parser.add_argument("--output-xlsx", dest="output_xlsx", default=None, help="Also write results to this Excel .xlsx file with boolean columns")
    parser.add_argument("--input-sheet", dest="input_sheet", default=None, help="XLSX sheet name to read (default: first sheet)")
    parser.add_argument("--blank-nontrue", action="store_true", help="When updating XLSX in-place, blank cells that are not True (default: leave as-is to preserve styles)")

    args = parser.parse_args(argv)

    ensure_bs4()

    # Debug mode: fetch a single URL and print Information text
    if args.debug_url:
        ensure_bs4()
        if not args.debug_url.startswith(APP_STORE_PREFIX):
            sys.stderr.write("--debug-url must be an Apple App Store URL (https://apps.apple.com...).\n")
            return 2
        session = make_session()
        html = load_page(args.debug_url, session, args.cache_dir, delay=args.delay)
        if not html:
            sys.stderr.write("Failed to load page.\n")
            return 1
        info_text = extract_information_text(html)
        print("=== Information section text ===")
        print(info_text)
        return 0

    # Debug mode: fetch a single URL and print detection flags + column mapping
    if args.debug_detect:
        if not args.debug_detect.startswith(APP_STORE_PREFIX):
            sys.stderr.write("--debug-detect must be an Apple App Store URL (https://apps.apple.com...).\n")
            return 2
        session = make_session()
        html = load_page(args.debug_detect, session, args.cache_dir, delay=args.delay)
        if not html:
            sys.stderr.write("Failed to load page.\n")
            return 1
        info_text = extract_information_text(html)
        flags = detect_compat(info_text)
        # Print flags and the resolved columns using current header mapping from a quick read of input (if available)
        try:
            rows_tmp, header_tmp, _ = read_csv_rows(args.input_csv)
        except Exception:
            header_tmp = []
        col_map_tmp = resolve_column_map(header_tmp)
        print("=== Detected flags ===")
        for k in sorted(flags.keys()):
            print(f"{k}: {flags[k]}")
        print("=== Column mapping ===")
        for k in sorted(col_map_tmp.keys()):
            print(f"{k} -> {col_map_tmp[k]}")
        return 0

    is_xlsx_input = args.input_csv.lower().endswith(".xlsx")
    if is_xlsx_input:
        rows, header = read_xlsx_rows(args.input_csv, sheet_name=args.input_sheet)
        delimiter = ","  # unused for XLSX
    else:
        rows, header, delimiter = read_csv_rows(args.input_csv)
    col_map = resolve_column_map(header)

    # Ensure target columns exist
    for col in col_map.values():
        if col not in header:
            header.append(col)

    url_col = args.url_col or detect_url_column(header, rows)
    if not url_col:
        sys.stderr.write("Could not detect the App Store URL column. Use --url-column to specify.\n")
        return 2

    session = make_session()

    processed = 0
    updated = 0

    def row_complete(r: Dict[str, str]) -> bool:
        # Consider complete only if all target columns are non-empty
        for col in col_map.values():
            v = (r.get(col) or "").strip()
            if v == "":
                return False
        return True

    for idx, r in enumerate(rows):
        url = (r.get(url_col) or "").strip()
        if not url.startswith(APP_STORE_PREFIX):
            continue
        if args.resume and row_complete(r):
            continue
        processed += 1
        if args.limit and processed > args.limit:
            break

        html = load_page(url, session, args.cache_dir, delay=args.delay)
        if not html:
            continue
        info_text = extract_information_text(html)
        flags = detect_compat(info_text)

    # No additional platform variant fetching; rely solely on primary page Information section

        changed = False
        for token, is_true in flags.items():
            out_col = col_map.get(token, COMPAT_COLUMN_MAP[token])
            if is_true and (r.get(out_col) or "").strip().lower() != "true":
                r[out_col] = "True"
                changed = True
        if changed:
            updated += 1

        if processed % 20 == 0:
            print(f"Processed {processed} App Store URLs... updated rows so far: {updated}")

    if is_xlsx_input:
        # For XLSX inputs, update the original workbook in-place style to preserve validations/formatting.
        xlsx_path = args.output_xlsx or re.sub(r"\.xlsx$", "_compat.xlsx", args.input_csv, flags=re.I)
        col_map2 = resolve_column_map(header)
        target_cols = [
            col_map2.get("ipad", "iPad"),
            col_map2.get("ipod touch", "iPod touch"),
            col_map2.get("apple tv", "Apple TV"),
            col_map2.get("mac", "Mac"),
            col_map2.get("apple watch", "Apple Watch"),
            col_map2.get("apple vision pro", "Apple Vision"),
            col_map2.get("iphone", "iPhone"),
        ]
        target_cols = [c for c in target_cols if c]
        updated_cells, missing_cols, effective_xlsx = update_xlsx_inplace(
            args.input_csv, xlsx_path, args.input_sheet, target_cols, rows, blank_nontrue=args.blank_nontrue
        )
        print(f"Done. Processed App Store rows: {processed}, updated rows: {updated}. Output: {effective_xlsx}")
        if missing_cols:
            print(f"Note: columns not found in sheet (skipped): {', '.join(missing_cols)}")
        # If CSV path explicitly provided, write CSV too
        if args.output_csv:
            write_csv_rows(args.output_csv, header, rows, ",")
            print(f"Also wrote CSV: {args.output_csv}")
    else:
        out_path = args.output_csv or re.sub(r"\.csv$", "_compat.csv", args.input_csv, flags=re.I)
        write_csv_rows(out_path, header, rows, delimiter)
        print(f"Done. Processed App Store rows: {processed}, updated rows: {updated}. Output: {out_path}")
        # For CSV inputs only: optional XLSX export with boolean coercion for selected columns
    if (not is_xlsx_input) and args.output_xlsx:
        col_map2 = resolve_column_map(header)
        bool_cols = [
            col_map2.get("ipad", "iPad"),
            col_map2.get("ipod touch", "iPod touch"),
            col_map2.get("apple tv", "Apple TV"),
            col_map2.get("mac", "Mac"),
            col_map2.get("apple watch", "Apple Watch"),
            col_map2.get("apple vision pro", "Apple Vision"),
            col_map2.get("iphone", "iPhone"),
        ]
        bool_cols = [c for c in bool_cols if c in header]
        xlsx_path = args.output_xlsx
        try:
            write_xlsx_rows(xlsx_path, header, rows, bool_cols)
            print(f"Also wrote Excel file: {xlsx_path}")
        except Exception as exc:
            print(f"Warning: could not write Excel file ({exc}).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
